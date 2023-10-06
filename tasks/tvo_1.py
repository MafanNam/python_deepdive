import openpyxl, math
from matplotlib import pyplot as plt

# Завантаження Excel-файлу та вибір активного аркуша
wb = openpyxl.load_workbook('DATA.xlsx')
ws = wb.active

# Список назв показників (з першого рядка таблиці)
names = [ws[1][i].value for i in range(0, 11)]

values = [[], [], [], [], [], [], [], [], [], [], []]

# Список квадратів відхилень від середнього
sum_column = [[], [], [], [], [], [], [], [], [], [], []]

for row in range(2, 2002):
    for column in range(11):
        values[column].append(float(ws[row][column].value))

# Обчислення середніх значень показників
average = list(map(lambda x: sum(x) / 2000, values))

# Обчислення квадратів відхилень від середнього значення
for i in range(len(values[0])):
    for j in range(11):
        sum_column[j].append((values[j][i] - average[j]) ** 2)

# Обчислення стандартних відхилень
for suma in range(len(sum_column)):
    sum_column[suma] = math.sqrt(sum(sum_column[suma]) / (len(sum_column[suma]) - 1))

for name, sum in list(zip(names, sum_column)):
    print(name, round(sum * 100, 2))

print("Найменше відхилення від середнього значення: Прод R " + str(min(sum_column)))

fig = plt.figure(figsize=(10, 7))
plt.bar(names, sum_column)
plt.title("Відхилення від середнього значення")
plt.xlabel("Показник")
plt.ylabel("Значення")
plt.show()
